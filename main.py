import openpyxl
import os
import tkinter as tk
from tkinter import messagebox
import winshell
from win32com.client import Dispatch


# Função para criar o arquivo Excel se ele não existir
def criar_arquivo_excel():
    if not os.path.exists('clientes.xlsx'):
        wb = openpyxl.Workbook()
        wb.save('clientes.xlsx')


# Função para adicionar um novo cliente ao arquivo Excel
def adicionar_cliente():
    nome = entry_nome.get()
    cpf = entry_cpf.get()
    endereco = entry_endereco.get()
    telefone = entry_telefone.get()
    valor_devido = float(entry_valor_devido.get())

    wb = openpyxl.load_workbook('clientes.xlsx')
    sheet = wb.active
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1).value = nome
    sheet.cell(row=next_row, column=2).value = cpf
    sheet.cell(row=next_row, column=3).value = endereco
    sheet.cell(row=next_row, column=4).value = telefone
    sheet.cell(row=next_row, column=5).value = valor_devido
    wb.save('clientes.xlsx')
    messagebox.showinfo("Sucesso", "Cliente adicionado com sucesso.")
    limpar_campos()


# Função para buscar o valor devido de um cliente
def buscar_valor_devido():
    nome = entry_nome_busca.get()
    wb = openpyxl.load_workbook('clientes.xlsx')
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == nome:
            valor_devido = sheet.cell(row=row, column=5).value
            messagebox.showinfo("Valor Devido", f"O cliente {nome} deve R${valor_devido}.")
            return
    messagebox.showinfo("Cliente não encontrado", "O cliente não foi encontrado.")


# Função para atualizar o valor pago de um cliente existente
def atualizar_valor_pago():
    nome = entry_nome_atualizar.get()
    novo_valor_pago = float(entry_novo_valor.get())

    wb = openpyxl.load_workbook('clientes.xlsx')
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == nome:
            valor_devido = sheet.cell(row=row, column=5).value
            valor_devido -= novo_valor_pago
            if valor_devido < 0:
                valor_devido = 0
            sheet.cell(row=row, column=5).value = valor_devido

            # Verificar se a célula não é None antes de adicionar o novo valor pago
            valor_pago_celula = sheet.cell(row=row, column=6).value
            if valor_pago_celula is None:
                valor_pago_celula = 0

            sheet.cell(row=row, column=6).value = valor_pago_celula + novo_valor_pago

            wb.save('clientes.xlsx')
            messagebox.showinfo("Sucesso", "Valor pago atualizado com sucesso.")
            return
    messagebox.showinfo("Cliente não encontrado", "O cliente não foi encontrado.")


# Função para adicionar novo débito ao débito existente de um cliente
def novo_debito():
    nome = entry_nome_debito.get()
    novo_debito = float(entry_novo_debito.get())

    wb = openpyxl.load_workbook('clientes.xlsx')
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == nome:
            valor_devido = sheet.cell(row=row, column=5).value
            valor_devido += novo_debito
            sheet.cell(row=row, column=5).value = valor_devido
            wb.save('clientes.xlsx')
            messagebox.showinfo("Sucesso", "Novo débito adicionado com sucesso.")
            return
    messagebox.showinfo("Cliente não encontrado", "O cliente não foi encontrado.")


# Função para limpar os campos do formulário
def limpar_campos():
    entry_nome.delete(0, tk.END)
    entry_cpf.delete(0, tk.END)
    entry_endereco.delete(0, tk.END)
    entry_telefone.delete(0, tk.END)
    entry_valor_devido.delete(0, tk.END)
    entry_nome_debito.delete(0, tk.END)
    entry_novo_debito.delete(0, tk.END)


# Função para criar a interface gráfica
def criar_interface():
    window = tk.Tk()
    window.title("Cadastro de Clientes")

    # Definir a janela para estar sempre no topo
    window.wm_attributes("-topmost", True)

    # Frame para adicionar cliente
    frame_adicionar = tk.LabelFrame(window, text="Adicionar Cliente")
    frame_adicionar.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)

    label_nome = tk.Label(frame_adicionar, text="Nome:")
    label_nome.grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)

    label_cpf = tk.Label(frame_adicionar, text="CPF:")
    label_cpf.grid(row=1, column=0, padx=5, pady=2, sticky=tk.W)

    label_endereco = tk.Label(frame_adicionar, text="Endereço:")
    label_endereco.grid(row=2, column=0, padx=5, pady=2, sticky=tk.W)

    label_telefone = tk.Label(frame_adicionar, text="Telefone:")
    label_telefone.grid(row=3, column=0, padx=5, pady=2, sticky=tk.W)

    label_valor_devido = tk.Label(frame_adicionar, text="Valor Devido:")
    label_valor_devido.grid(row=4, column=0, padx=5, pady=2, sticky=tk.W)

    global entry_nome, entry_cpf, entry_endereco, entry_telefone, entry_valor_devido
    entry_nome = tk.Entry(frame_adicionar)
    entry_nome.grid(row=0, column=1, padx=5, pady=2)

    entry_cpf = tk.Entry(frame_adicionar)
    entry_cpf.grid(row=1, column=1, padx=5, pady=2)

    entry_endereco = tk.Entry(frame_adicionar)
    entry_endereco.grid(row=2, column=1, padx=5, pady=2)

    entry_telefone = tk.Entry(frame_adicionar)
    entry_telefone.grid(row=3, column=1, padx=5, pady=2)

    entry_valor_devido = tk.Entry(frame_adicionar)
    entry_valor_devido.grid(row=4, column=1, padx=5, pady=2)

    btn_adicionar = tk.Button(frame_adicionar, text="Adicionar Cliente", command=adicionar_cliente)
    btn_adicionar.grid(row=5, column=0, columnspan=2, padx=5, pady=5)

    # Frame para buscar cliente
    frame_buscar = tk.LabelFrame(window, text="Buscar Cliente")
    frame_buscar.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)

    label_nome_busca = tk.Label(frame_buscar, text="Nome:")
    label_nome_busca.grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)

    global entry_nome_busca
    entry_nome_busca = tk.Entry(frame_buscar)
    entry_nome_busca.grid(row=0, column=1, padx=5, pady=2)

    btn_buscar = tk.Button(frame_buscar, text="Buscar", command=buscar_valor_devido)
    btn_buscar.grid(row=1, column=0, columnspan=2, padx=5, pady=5)

    # Frame para atualizar valor pago
    frame_atualizar = tk.LabelFrame(window, text="Atualizar Valor Pago")
    frame_atualizar.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)

    label_nome_atualizar = tk.Label(frame_atualizar, text="Nome:")
    label_nome_atualizar.grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)

    label_novo_valor = tk.Label(frame_atualizar, text="Novo Valor Pago:")
    label_novo_valor.grid(row=1, column=0, padx=5, pady=2, sticky=tk.W)

    global entry_nome_atualizar, entry_novo_valor
    entry_nome_atualizar = tk.Entry(frame_atualizar)
    entry_nome_atualizar.grid(row=0, column=1, padx=5, pady=2)

    entry_novo_valor = tk.Entry(frame_atualizar)
    entry_novo_valor.grid(row=1, column=1, padx=5, pady=2)

    btn_atualizar = tk.Button(frame_atualizar, text="Atualizar Valor Pago", command=atualizar_valor_pago)
    btn_atualizar.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

    # Frame para adicionar novo débito
    frame_debito = tk.LabelFrame(window, text="Adicionar Novo Débito")
    frame_debito.grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)

    label_nome_debito = tk.Label(frame_debito, text="Nome:")
    label_nome_debito.grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)

    label_novo_debito = tk.Label(frame_debito, text="Novo Débito:")
    label_novo_debito.grid(row=1, column=0, padx=5, pady=2, sticky=tk.W)

    global entry_nome_debito, entry_novo_debito
    entry_nome_debito = tk.Entry(frame_debito)
    entry_nome_debito.grid(row=0, column=1, padx=5, pady=2)

    entry_novo_debito = tk.Entry(frame_debito)
    entry_novo_debito.grid(row=1, column=1, padx=5, pady=2)

    btn_debito = tk.Button(frame_debito, text="Adicionar Débito", command=novo_debito)
    btn_debito.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

    window.mainloop()


# Função para criar o atalho na área de trabalho
def criar_atalho(nome_atalho, caminho_executavel):
    desktop = winshell.desktop()
    path = os.path.join(desktop, f"{nome_atalho}.lnk")
    target = caminho_executavel
    wDir = os.path.dirname(target)
    icon = target
    shell = Dispatch('WScript.Shell')
    shortcut = shell.CreateShortCut(path)
    shortcut.Targetpath = target
    shortcut.WorkingDirectory = wDir
    shortcut.IconLocation = icon
    shortcut.save()


# Caminho completo para o executável do Python
caminho_python = r"C:\Usuários\mf.joa\Python-Dede/cadastro_clientes.py"  # Substitua pelo caminho real do executável do Python

# Nome do atalho
nome_atalho = "Cadastro de Clientes"

# Criar o atalho na área de trabalho
criar_atalho(nome_atalho, caminho_python)

# Início do programa
criar_arquivo_excel()
criar_interface()
